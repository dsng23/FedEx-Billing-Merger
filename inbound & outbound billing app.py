import io
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side
import streamlit as st

# Website setup
st.set_page_config(page_title="FedEx Billing Merger", page_icon="📦", layout="wide")

st.markdown("""
    <style>
    html, body, [class*="css"] { font-size: 18px !important; }
    .fedex-title { color: #4D148C; font-weight: 900; font-size: 48px; letter-spacing: -1px; margin-bottom: 5px; }
    .fedex-divider { height: 6px; background: linear-gradient(to right, #4D148C 50%, #FF6600 50%); margin-bottom: 30px; border-radius: 3px; }
    .instruction-box { background-color: #F8F9FA; border-left: 5px solid #FF6600; padding: 20px; border-radius: 5px; margin-bottom: 30px; font-size: 18px; color: #333; }
    </style>
""", unsafe_allow_html=True)

st.markdown('<div class="fedex-title">📦 FedEx Billing Merger</div>', unsafe_allow_html=True)
st.markdown('<div class="fedex-divider"></div>', unsafe_allow_html=True)

# Merging either inbound or outbound files
workflow = st.radio("🔀 Select your workflow:", ["Inbound", "Outbound"], horizontal=True)

st.write("") 

# OUTBOUND WORKFLOW
if workflow == "Outbound":
    
    # --- 2. UPLOAD FILES ---
    st.markdown("""
    <div class="instruction-box">
        <strong>Outbound File Portal</strong><br> 
        Upload the two Outbound Excel files below. Use the "Rows to skip" to bypass the junk titles. 
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**File 1 (Outbound Header Data)**")
        file1 = st.file_uploader("📂 Upload Outbound Header File", type=["xlsx", "xls", "csv"], key="out_f1")
        skip_rows_1 = st.number_input("⚙️ Rows to skip at top of File 1:", min_value=0, value=0, key="out_skip1")

    with col2:
        st.markdown("**File 2 (Outbound Body Data)**")
        file2 = st.file_uploader("📂 Upload Outbound Body File", type=["xlsx", "xls", "csv"], key="out_f2")
        skip_rows_2 = st.number_input("⚙️ Rows to skip at top of File 2:", min_value=0, value=0, key="out_skip2")

    with st.expander("❓ What does 'Rows to skip' mean?"):
        st.write("""
        **Rows to skip"** are the number of extra lines at the top of Excel files before real data starts (like report titles, company logos, or blank spaces). **Example:** if headers start on row 8, then the number of junk rows = 7.
        """)

    st.divider()
    st.subheader("⚙️ Step 2: Output Settings")
    custom_filename = st.text_input("📝 Name your final Excel file:", value="Outbound_Merged_Report", key="out_name")
    
    final_filename = custom_filename.strip()
    if not final_filename.endswith(".xlsx"):
        final_filename += ".xlsx"
    st.write("") 

    # Only run if both files are uploaded
    if file1 is not None and file2 is not None:
        if st.button("🚀 Process & Combine Outbound Files", type="primary", key="out_btn"):
            try:
                # --- 3. READ AND CLEAN RAW DATA ---
                # Load files into the system
                if file1.name.endswith('.csv'):
                    df1 = pd.read_csv(file1, skiprows=skip_rows_1)
                else:
                    df1 = pd.read_excel(file1, skiprows=skip_rows_1)
                    
                if file2.name.endswith('.csv'):
                    df2 = pd.read_csv(file2, skiprows=skip_rows_2)
                else:
                    df2 = pd.read_excel(file2, skiprows=skip_rows_2)
                
                # Fix blank spaces in column names
                df1.columns = df1.columns.astype(str).str.strip()
                df2.columns = df2.columns.astype(str).str.strip()
                df1 = df1.replace(r'^\s*$', np.nan, regex=True).replace('', np.nan)
                df2 = df2.replace(r'^\s*$', np.nan, regex=True).replace('', np.nan)

                # Rename raw system columns to match the final report format
                outbound_mapping = {
                    'SHIP DATE': 'Date Out', 'REF NBR': 'Inv No.', 'SHIP TO NAME': 'Cnee',
                    'SHIP TO COUNTRY': 'Dest', 'ORDERED QTY': 'Qty', 'LPN TYPE': 'Mat Type',
                    'ORDER': 'Order Nbr', 'ORDER NBR': 'Order Nbr', 'ITEM': 'Item Code',
                    'ITEMS': 'Header Items', 'AWB/CONSIGNMENT NUMBER': 'Con Note'
                }
                df1.columns = [outbound_mapping.get(str(col).strip().upper(), str(col).strip()) for col in df1.columns]
                df2.columns = [outbound_mapping.get(str(col).strip().upper(), str(col).strip()) for col in df2.columns]

                # Format Order Numbers so they match perfectly across both files
                if 'Order Nbr' in df1.columns:
                    df1['Order Nbr'] = df1['Order Nbr'].astype(str).str.strip().str.upper().str.lstrip('0')
                    df1 = df1[~df1['Order Nbr'].isin(['NAN', 'NONE', ''])]
                if 'Order Nbr' in df2.columns:
                    df2['Order Nbr'] = df2['Order Nbr'].astype(str).str.strip().str.upper().str.lstrip('0')
                    df2 = df2[~df2['Order Nbr'].isin(['NAN', 'NONE', ''])]

                # Keep only the columns we actually need
                target_cols_1 = ['Date Out', 'Inv No.', 'Cnee', 'Dest', 'Order Nbr', 'Header Items', 'Con Note']
                target_cols_2 = ['Qty', 'Mat Type', 'Order Nbr', 'Item Code']
                df1_clean = df1[[c for c in df1.columns if c in target_cols_1]].copy()
                df2_clean = df2[[c for c in df2.columns if c in target_cols_2]].copy()

                df1_clean = df1_clean.loc[:, ~df1_clean.columns.duplicated()]
                df2_clean = df2_clean.loc[:, ~df2_clean.columns.duplicated()]

                # Split orders grouped by hyphens (e.g. Order1-Order2) into separate rows
                if 'Order Nbr' in df1_clean.columns:
                    df1_clean['Order Nbr'] = df1_clean['Order Nbr'].str.split('-')
                    df1_clean = df1_clean.explode('Order Nbr')
                    df1_clean['Order Nbr'] = df1_clean['Order Nbr'].str.strip()
                    df1_clean = df1_clean.drop_duplicates(subset=['Order Nbr'], keep='first')
                
                # Remove junk rows that don't have a quantity
                if 'Qty' in df2_clean.columns:
                    df2_clean = df2_clean.dropna(subset=['Qty'])

                # --- 4. MERGE THE TWO FILES ---
                # Link files using the Order Number
                if 'Order Nbr' in df1_clean.columns and 'Order Nbr' in df2_clean.columns:
                    overlap = set(df1_clean.columns).intersection(set(df2_clean.columns)) - {'Order Nbr'}
                    for col in overlap:
                        df2_clean = df2_clean.drop(columns=[col])
                    combined_df = pd.merge(df1_clean, df2_clean, on='Order Nbr', how='outer', indicator=True)
                else:
                    st.warning("⚠️ Could not find 'Order Nbr'/'Order' in both files. The system stacked them instead of linking them.")
                    combined_df = pd.concat([df1_clean, df2_clean], ignore_index=True)
                    combined_df['_merge'] = 'both'

                # Backup Link: If Order Number fails, try linking them using the Item Code instead
                if 'Header Items' in df1_clean.columns and 'Item Code' in combined_df.columns:
                    unmatched_mask = combined_df['_merge'] == 'right_only'
                    unmatched_body = combined_df[unmatched_mask].copy()
                    matched_body = combined_df[~unmatched_mask].copy()
                    
                    if not unmatched_body.empty:
                        rescue_df = df1_clean.dropna(subset=['Header Items']).copy()
                        rescue_df['Item Code'] = rescue_df['Header Items'].astype(str).str.split()
                        rescue_df = rescue_df.explode('Item Code')
                        rescue_df['Item Code'] = rescue_df['Item Code'].str.strip().str.upper()
                        rescue_df = rescue_df.drop_duplicates(subset=['Item Code'], keep='first')
                        
                        cols_to_fill = ['Date Out', 'Inv No.', 'Cnee', 'Dest', 'Con Note']
                        existing_fill_cols = [c for c in cols_to_fill if c in unmatched_body.columns]
                        unmatched_body = unmatched_body.drop(columns=existing_fill_cols + ['Header Items'], errors='ignore')
                        rescue_cols = ['Item Code'] + [c for c in cols_to_fill if c in rescue_df.columns]
                        
                        rescued_df = pd.merge(unmatched_body, rescue_df[rescue_cols], on='Item Code', how='left')
                        rescued_df['_merge'] = np.where(rescued_df['Inv No.'].notna(), 'rescued', 'right_only')
                        combined_df = pd.concat([matched_body, rescued_df], ignore_index=True)

                # Find items that completely failed to match and log them
                unmatched_df = combined_df[~combined_df['_merge'].isin(['both', 'rescued'])].copy()
                if not unmatched_df.empty:
                    unmatched_df['Error Reason'] = unmatched_df['_merge'].map({
                        'left_only': 'Header invoice found, but no matching products found in Body.',
                        'right_only': 'Products found in Body, but no matching Header OR Item Code found.'
                    })
                    unmatched_df = unmatched_df.drop(columns=['_merge'], errors='ignore')
                    log_cols = [c for c in unmatched_df.columns.tolist() if c != 'Error Reason']
                    front_cols = [c for c in ['Order Nbr', 'Item Code', 'Qty'] if c in log_cols]
                    back_cols = [c for c in log_cols if c not in front_cols]
                    unmatched_df = unmatched_df[front_cols + back_cols + ['Error Reason']]

                # Keep only successful matches for the final report
                combined_df = combined_df[combined_df['_merge'].isin(['both', 'rescued'])].copy()
                combined_df = combined_df.drop(columns=['_merge', 'Header Items', 'Item Code'], errors='ignore')

                # Define the final column order for the report
                final_column_order = ['Date Out', 'Con Note', 'Inv No.', 'Cnee', 'Dest', 'Temp', 'Qty', 'Mat Type']
                for col in final_column_order:
                    if col not in combined_df.columns:
                        combined_df[col] = ""
                combined_df = combined_df[final_column_order]

                # Sort by Invoice Number
                if 'Inv No.' in combined_df.columns:
                    combined_df = combined_df.sort_values(['Inv No.', 'Date Out'], ascending=[True, True], na_position='last').reset_index(drop=True)

                # --- 5. SHOW LIVE PREVIEW ON WEBSITE ---
                st.markdown('<div class="fedex-divider"></div>', unsafe_allow_html=True)
                st.subheader("✏️ Review and Edit Outbound Data")
                st.write("📝 Click directly into any cell to fix data. Click a column header to sort rows, or hover on the left to add/delete rows.")
                edited_df = st.data_editor(combined_df, use_container_width=True, num_rows="dynamic")
                st.write("") 

                # --- 6. BUILD THE FINAL EXCEL DOWNLOAD FILE ---
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    
                    # Create main report sheet
                    edited_df.to_excel(writer, index=False, sheet_name='Outbound Report')
                    worksheet = writer.sheets['Outbound Report']

                    # Setup purple header design and borders
                    header_fill = PatternFill(start_color="4D148C", end_color="4D148C", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF", size=14)
                    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    for cell in worksheet[1]:
                        cell.fill, cell.font, cell.border = header_fill, header_font, thick_border
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                        for cell in row: cell.border = thin_border

                    # Resize columns to fit text
                    for idx, col in enumerate(edited_df.columns):
                        max_len = max(int(edited_df[col].astype(str).str.len().max() if pd.notna(edited_df[col].astype(str).str.len().max()) else 0), int(len(str(col)) * 1.4))
                        worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len + 4
                    worksheet.freeze_panes = 'A2'

                    # Build Summary Table 1: Quantity by Destination
                    summary_df = edited_df.copy()
                    summary_df['Qty_num'] = pd.to_numeric(summary_df['Qty'], errors='coerce').fillna(0)
                    summary_df['Dest'] = summary_df['Dest'].replace(r'^\s*$', '(Blank)', regex=True).fillna('(Blank)')
                    summary_df['Inv No.'] = summary_df['Inv No.'].replace(r'^\s*$', '(Blank)', regex=True).fillna('(Blank)')
                    outbound_sums = summary_df.groupby(['Dest', 'Inv No.'])['Qty_num'].sum().reset_index()

                    worksheet.merge_cells('J4:L4')
                    title_cell = worksheet.cell(row=4, column=10, value="SUM OF QUANTITY")
                    title_cell.fill, title_cell.font, title_cell.border = header_fill, header_font, thick_border
                    worksheet.cell(row=4, column=11).border, worksheet.cell(row=4, column=12).border = thick_border, thick_border 

                    for col_idx, val in enumerate(["Dest", "Inv No.", "Total"], start=10):
                        cell = worksheet.cell(row=5, column=col_idx, value=val)
                        cell.font, cell.border = Font(bold=True), thin_border

                    current_row, subtotal_rows = 6, []
                    for dest, group in outbound_sums.groupby('Dest', sort=True):
                        start_dest_row = current_row
                        for _, row_data in group.sort_values(by='Inv No.').iterrows():
                            worksheet.cell(row=current_row, column=10, value=row_data['Dest']).border = thin_border
                            worksheet.cell(row=current_row, column=11, value=row_data['Inv No.']).border = thin_border
                            worksheet.cell(row=current_row, column=12, value=row_data['Qty_num']).border = thin_border
                            current_row += 1
                            
                        sub_v = worksheet.cell(row=current_row, column=10, value=f"{dest} Total")
                        worksheet.cell(row=current_row, column=11, value="").border = thin_border 
                        sub_formula = f"=SUM(L{start_dest_row}:L{current_row - 1})" if current_row > start_dest_row else "=0"
                        sub_t = worksheet.cell(row=current_row, column=12, value=sub_formula)
                        sub_v.font, sub_t.font, sub_v.border, sub_t.border = Font(bold=True), Font(bold=True), thin_border, thin_border
                        subtotal_rows.append(f"L{current_row}")
                        current_row += 1

                    gt_v = worksheet.cell(row=current_row, column=10, value="Grand Total")
                    worksheet.cell(row=current_row, column=11, value="").border = thin_border 
                    gt_formula = "=" + "+".join(subtotal_rows) if subtotal_rows else "=0"
                    gt_t = worksheet.cell(row=current_row, column=12, value=gt_formula)
                    gt_v.font, gt_t.font, gt_v.border, gt_t.border = Font(bold=True), Font(bold=True), thin_border, thin_border
                    worksheet.column_dimensions['J'].width, worksheet.column_dimensions['K'].width, worksheet.column_dimensions['L'].width = 15, 25, 15

                    # Build Summary Table 2: Quantity by Temperature
                    t2_start_row = current_row + 2
                    worksheet.merge_cells(start_row=t2_start_row, start_column=11, end_row=t2_start_row, end_column=12)
                    t2_title = worksheet.cell(row=t2_start_row, column=11, value="Total OUT")
                    t2_title.fill, t2_title.font, t2_title.border = header_fill, header_font, thick_border
                    worksheet.cell(row=t2_start_row, column=12).border = thick_border 
                    
                    for col_idx, val in enumerate(["Temp", "Total"], start=11):
                        cell = worksheet.cell(row=t2_start_row + 1, column=col_idx, value=val)
                        cell.font, cell.border = Font(bold=True), thin_border

                    t2_current_row = t2_start_row + 2
                    max_data_row = max(len(edited_df) + 1, 2)
                    
                    for temp_cat in ["Ambient", "Chilled", "Frozen"]:
                        worksheet.cell(row=t2_current_row, column=11, value=temp_cat).border = thin_border
                        worksheet.cell(row=t2_current_row, column=12, value=f'=SUMIFS(G$2:G${max_data_row}, F$2:F${max_data_row}, K{t2_current_row})').border = thin_border
                        t2_current_row += 1

                    # Add Unmatched Log as a second Excel sheet
                    if not unmatched_df.empty:
                        unmatched_df.to_excel(writer, index=False, sheet_name='Unmatched Log')
                        ws_unmatched = writer.sheets['Unmatched Log']
                        for cell in ws_unmatched[1]:
                            cell.fill, cell.font = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"), Font(bold=True, color="FFFFFF")
                        for idx, col in enumerate(unmatched_df.columns):
                            max_len = max(15, int(unmatched_df[col].astype(str).str.len().max() if pd.notna(unmatched_df[col].astype(str).str.len().max()) else 15))
                            ws_unmatched.column_dimensions[get_column_letter(idx + 1)].width = max_len + 2

                # Rewind the invisible Excel file back to the beginning so the download button works
                buffer.seek(0)
                st.success(f"🎉 Your outbound files were successfully processed! Ready for download.")
                if not unmatched_df.empty:
                    st.warning(f"⚠️ We found {len(unmatched_df)} rows that failed to match. They have been removed from the main report and logged in the 'Unmatched Log' tab!")
                
                st.download_button(label=f"📥 Download {final_filename}", data=buffer, file_name=final_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="out_dl")

            except Exception as e:
                st.error("🚨 An error occurred while parsing the files.")
                st.exception(e)
    else:
        st.info("Waiting for both Outbound files to be uploaded...")

# INBOUND WORKFLOW 
if workflow == "Inbound":

    # Show instructions and create the upload boxes for the two files
    st.markdown("""
    <div class="instruction-box">
        <strong>Inbound File Portal</strong><br> 
        Upload the two Inbound Excel files below. Use the "Rows to skip" to bypass the junk titles. 
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**File 1 (Base Data)**")
        file1 = st.file_uploader("📂 Upload Inbound Dock Date File", type=["xlsx", "xls", "csv"], key="in_f1")
        skip_rows_1 = st.number_input("⚙️ Rows to skip at top of File 1:", min_value=0, value=0, key="in_skip1")

    with col2:
        st.markdown("**File 2 (Data to Merge)**")
        file2 = st.file_uploader("📂 Upload Inbound Shipment File", type=["xlsx", "xls", "csv"], key="in_f2")
        skip_rows_2 = st.number_input("⚙️ Rows to skip at top of File 2:", min_value=0, value=0, key="in_skip2")

    # Explanation dropdown for the user
    with st.expander("❓ What does 'Rows to skip' mean?"):
        st.write("""
        **Rows to skip"** are the number of extra lines at the top of Excel files before real data starts (like report titles, company logos, or blank spaces). **Example:** if headers start on row 8, then the number of junk rows = 7.
        """)

    st.divider() 
    
    # Name the final downloaded report
    st.subheader("⚙️ Step 2: Output Settings")
    custom_filename = st.text_input("📝 Name your final Excel file:", value="Inbound_Merged_Report", key="in_name")
    
    final_filename = custom_filename.strip()
    if not final_filename.endswith(".xlsx"):
        final_filename += ".xlsx"
        
    st.write("")

    # Only run code if both files are successfully uploaded
    if file1 is not None and file2 is not None:
    
        if st.button("🚀 Process & Combine Files", type="primary", key="in_btn"):
            try:
                # Read the files into the system
                if file1.name.endswith('.csv'):
                    df1 = pd.read_csv(file1, skiprows=skip_rows_1)
                else:
                    df1 = pd.read_excel(file1, skiprows=skip_rows_1)

                if file2.name.endswith('.csv'):
                    df2 = pd.read_csv(file2, skiprows=skip_rows_2)
                else:
                    df2 = pd.read_excel(file2, skiprows=skip_rows_2)

                # Clean up accidental spaces in column names
                df1.columns = df1.columns.astype(str).str.strip()
                df2.columns = df2.columns.astype(str).str.strip()

                # Find empty spaces and fill them in with data from the row above (handles merged Excel cells)
                df1 = df1.replace(r'^\s*$', np.nan, regex=True).replace('', np.nan).ffill()
                df2 = df2.replace(r'^\s*$', np.nan, regex=True).replace('', np.nan).ffill()

                # Rename the raw system columns into the clean final report column names
                hardcoded_mapping = {
                    'DOCK DATE & TIME': 'Date Time IN',
                    'MANIFEST NBR': 'Airwaybill No.',
                    'LINE REFERENCE': 'PO Number',
                    'CUST REF NUM': 'Invoice No.',
                    'VENDOR INFO': 'Vendor Name',
                    'ITEM CODE': 'Supplier P/N',
                    'SECONDARY ITEM CODE': 'SAP No.',
                    'ITEM DESCRIPTION': 'Description',
                    'STORAGE TEMPERATURE': 'Temp',
                    'SHIPPED QTY': 'Qty',
                    'SHIPMENT NBR': 'SHIPMENT NBR' 
                }
                
                df1.columns = [hardcoded_mapping.get(str(col).strip().upper(), str(col).strip()) for col in df1.columns]
                df2.columns = [hardcoded_mapping.get(str(col).strip().upper(), str(col).strip()) for col in df2.columns]

                # Match and link the two files together using the Shipment Number
                if 'SHIPMENT NBR' in df1.columns and 'SHIPMENT NBR' in df2.columns:
                    combined_df = pd.merge(df1, df2, on='SHIPMENT NBR', how='outer')

                    for col in list(hardcoded_mapping.values()):
                        col_x = f"{col}_x"
                        col_y = f"{col}_y"
                        if col_x in combined_df.columns and col_y in combined_df.columns:
                            combined_df[col] = combined_df[col_x].combine_first(combined_df[col_y])
                            combined_df = combined_df.drop(columns=[col_x, col_y])
                else:
                    st.warning("⚠️ Could not find 'SHIPMENT NBR' in both files. Make sure the number of rows to skip is correct for each file.")
                    combined_df = pd.concat([df1, df2], ignore_index=True)

                # Remove junk 'Unnamed' columns generated by Excel exports
                combined_df = combined_df.loc[:, ~combined_df.columns.str.contains('^Unnamed')]

                # This list controls the exact left-to-right sequence of columns in the final Excel Report
                final_column_order = [
                    'Date Time IN', 'Airwaybill No.', 'PO Number', 'Invoice No.',
                    'Vendor Name', 'Supplier P/N', 'SAP No.', 'Description',
                    'Type of Goods', 'Temp', 'Qty'
                ]
            
                # Add any missing columns as blanks so the formatting remains consistent
                for col in final_column_order:
                    if col not in combined_df.columns:
                        combined_df[col] = ""
                        
                combined_df = combined_df[final_column_order]

                # Clean the formatting of the Temperature column so it groups nicely
                if 'Temp' in combined_df.columns:
                    combined_df['Temp'] = combined_df['Temp'].replace(np.nan, '')
                    combined_df['Temp'] = combined_df['Temp'].astype(str).str.strip().str.title()
                    combined_df['Temp'] = combined_df['Temp'].replace(['Dg Ambient', 'Nan'], ['Ambient', ''])

                # Show a live preview of the data on the website for quick editing
                st.markdown('<div class="fedex-divider"></div>', unsafe_allow_html=True)
                
                st.subheader("✏️ Review and Edit Data")
                st.write("📝 Click directly into any cell to fix data. Click a column header to sort rows, or hover on the left to add/delete rows.")
                
                edited_df = st.data_editor(
                    combined_df, 
                    use_container_width=True,
                    num_rows="dynamic" 
                )
                st.write("") 

                # Prepare the data to be written into an excel file
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    
                    # Create the main data sheet
                    edited_df.to_excel(writer, index=False, sheet_name='Master Report')
                    worksheet = writer.sheets['Master Report']

                    # Define the Excel header background color and text font
                    header_fill = PatternFill(start_color="4D148C", end_color="4D148C", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF", size=14)
                    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), 
                                          top=Side(style='thick'), bottom=Side(style='thick'))
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                         top=Side(style='thin'), bottom=Side(style='thin'))

                    # Apply formatting to headers
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = thick_border

                    # Apply thin borders to all data rows
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                        for cell in row:
                            cell.border = thin_border

                    # Automatically resize columns so the text isn't cut off
                    for idx, col in enumerate(edited_df.columns):
                        max_data_length = edited_df[col].astype(str).str.len().max()
                        max_data_length = 0 if pd.isna(max_data_length) else int(max_data_length)
                        header_length = int(len(str(col)) * 1.4)

                        perfect_width = max(max_data_length, header_length) + 4
                        col_letter = get_column_letter(idx + 1)
                        worksheet.column_dimensions[col_letter].width = perfect_width

                    worksheet.freeze_panes = 'A2'

                    # Build the Quantity grouped by Vendor Name Summary Table
                    summary_df = edited_df.copy()
                    summary_df['Qty_num'] = pd.to_numeric(summary_df['Qty'], errors='coerce').fillna(0)
                    summary_df['Vendor Name'] = summary_df['Vendor Name'].replace(r'^\s*$', '(Blank)', regex=True).fillna('(Blank)')
                    
                    vendor_sums = summary_df.groupby('Vendor Name')['Qty_num'].sum().reset_index()

                    worksheet.merge_cells('M4:N4')
                    title_cell = worksheet.cell(row=4, column=13, value="SUM OF QUANTITY")
                    title_cell.fill = header_fill
                    title_cell.font = header_font
                    title_cell.border = thick_border
                    worksheet.cell(row=4, column=14).border = thick_border

                    head_v = worksheet.cell(row=5, column=13, value="Vendor Name")
                    head_t = worksheet.cell(row=5, column=14, value="Total")
                    head_v.font = Font(bold=True)
                    head_t.font = Font(bold=True)
                    head_v.border = thin_border
                    head_t.border = thin_border
                
                    current_row = 6
                    start_data_row = 6
                    for _, row_data in vendor_sums.iterrows():
                        worksheet.cell(row=current_row, column=13, value=row_data['Vendor Name']).border = thin_border
                        worksheet.cell(row=current_row, column=14, value=row_data['Qty_num']).border = thin_border
                        current_row += 1

                    # Add final Grand Total row for the Quantity grouped by Vendor Name Summary Table
                    gt_v = worksheet.cell(row=current_row, column=13, value="Grand Total")
                    
                    if current_row > start_data_row:
                        sum_formula = f"=SUM(N{start_data_row}:N{current_row - 1})" 
                    else:
                        sum_formula = "=0"

                    gt_t = worksheet.cell(row=current_row, column=14, value=sum_formula)
                    
                    gt_v.font = Font(bold=True)
                    gt_t.font = Font(bold=True)
                    gt_v.border = thin_border
                    gt_t.border = thin_border

                    worksheet.column_dimensions['M'].width = 30
                    worksheet.column_dimensions['N'].width = 15
                    
                    # Build the Quantity grouped by Temperature Summary Table
                    t2_start_row = current_row + 2
                    
                    worksheet.merge_cells(start_row=t2_start_row, start_column=13, end_row=t2_start_row, end_column=14)
                    t2_title = worksheet.cell(row=t2_start_row, column=13, value="Total IN")
                    t2_title.fill = header_fill
                    t2_title.font = header_font
                    t2_title.border = thick_border
                    worksheet.cell(row=t2_start_row, column=14).border = thick_border 
                    
                    t2_head_v = worksheet.cell(row=t2_start_row + 1, column=13, value="Temp")
                    t2_head_t = worksheet.cell(row=t2_start_row + 1, column=14, value="Type")
                    t2_head_v.font = Font(bold=True)
                    t2_head_t.font = Font(bold=True)
                    t2_head_v.border = thin_border
                    t2_head_t.border = thin_border

                    temp_categories = ["Ambient", "Chilled", "Frozen"]
                    
                    t2_current_row = t2_start_row + 2
                    t2_start_data = t2_current_row
                    
                    max_data_row = len(edited_df) + 1
                    if max_data_row < 2:
                        max_data_row = 2
                    
                    for temp_cat in temp_categories:
                        worksheet.cell(row=t2_current_row, column=13, value=temp_cat).border = thin_border

                        sumifs_formula = f'=SUMIFS(K$2:K${max_data_row}, J$2:J${max_data_row}, M{t2_current_row})'
                        worksheet.cell(row=t2_current_row, column=14, value=sumifs_formula).border = thin_border
                        
                        t2_current_row += 1

                    # Add final Grand Total row for the Quantity grouped by Temperature Summary Table
                    t2_gt_v = worksheet.cell(row=t2_current_row, column=13, value="Grand Total")
                    t2_sum_formula = f"=SUM(N{t2_start_data}:N{t2_current_row - 1})"
                    t2_gt_t = worksheet.cell(row=t2_current_row, column=14, value=t2_sum_formula)
                    
                    t2_gt_v.font = Font(bold=True)
                    t2_gt_t.font = Font(bold=True) 
                    t2_gt_v.border = thin_border
                    t2_gt_t.border = thin_border

                buffer.seek(0)

                st.success(f"🎉 Your files were successfully processed, linked, and formatted! Ready for download.")
                
                # Renaming final merged file and ready for download
                st.download_button(
                    label=f"📥 Download {final_filename}",
                    data=buffer,
                    file_name=final_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="in_dl"
                )
            except Exception as e:
                st.error("🚨 An error occurred while parsing the files.")
                st.exception(e) 

    else:
        st.info("Waiting for both Inbound files to be uploaded...")

# ENTER: python -m streamlit run "c:/Users/1699678/OneDrive - MyFedEx/Desktop/billing app final ver.py" 
